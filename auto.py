import sys
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
from urllib.parse import quote
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException

def find_last_row_with_data(sheet):
    """
    Tìm hàng cuối cùng có dữ liệu thực sự trong sheet.
    """
    for row in range(sheet.max_row, 0, -1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value is not None:
                return row
    return 0

def get_chromedriver_path():
    """
    Trả về đường dẫn đến chromedriver.exe
    """
    if getattr(sys, 'frozen', False):
        # Đang chạy trong một .exe đã đóng gói bởi PyInstaller
        return os.path.join(sys._MEIPASS, 'chromedriver.exe')
    else:
        # Đang chạy trong môi trường phát triển
        return 'chromedriver.exe'

def main():
    # Tính ngày hôm qua
    yesterday = datetime.now() - timedelta(days=1)
    formatted_date = yesterday.strftime('%d/%m/%Y')
    formatted_yesterday = yesterday.strftime('%d/%m/%Y')  # Chuỗi định dạng ngày hôm qua

    # Mã hóa ngày hôm qua để thêm vào URL
    encoded_date = quote(formatted_date, safe='')

    
    # Tạo tùy chọn cho ChromeDriver
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Chạy chế độ ẩn
    chrome_options.add_argument("--disable-gpu")  # Tắt GPU (thường dùng trên Windows)
    chrome_options.add_argument("--window-size=1920x1080")  # Đặt kích thước cửa sổ (tùy chọn)
    chrome_options.add_argument("--disable-extensions")  # Tắt các tiện ích mở rộng
    chrome_options.add_argument("--no-sandbox")  # Thêm cờ này nếu chạy trên Linux
    chrome_options.add_argument("--log-level=3")

    # Khởi tạo WebDriver 
    chromedriver_path = get_chromedriver_path()
    try:
        driver = webdriver.Chrome(executable_path=chromedriver_path, options=chrome_options)
    except WebDriverException as e:
        print(f"Có lỗi trong quá trình khởi tạo WebDriver: {e}")
        return webdriver.Chrome(executable_path=chromedriver_path, options=chrome_options)  

    try:
        # Tạo URL với ngày hôm qua
        url_1cuasotk = f"http://motcuasothongke.cantho.gov.vn/group/guest/bao-cao-thong-ke?p_p_id=BaoCaoThongKe_WAR_ctonegateportlet&p_p_lifecycle=1&p_p_state=normal&p_p_mode=view&p_p_col_id=column-2&p_p_col_count=1&_BaoCaoThongKe_WAR_ctonegateportlet_hoSoLienThong=false&_BaoCaoThongKe_WAR_ctonegateportlet_isCongXuLyTon=1&_BaoCaoThongKe_WAR_ctonegateportlet_denNgay={encoded_date}&_BaoCaoThongKe_WAR_ctonegateportlet_tuNgay={encoded_date}&_BaoCaoThongKe_WAR_ctonegateportlet_maLinhVucJS=G15-TP10&_BaoCaoThongKe_WAR_ctonegateportlet_baiBo=-1&_BaoCaoThongKe_WAR_ctonegateportlet_javax.portlet.action=getChiTietThongKeTHCLHS&_BaoCaoThongKe_WAR_ctonegateportlet_maLinhVuc=G15-TP10&_BaoCaoThongKe_WAR_ctonegateportlet_linhVucId=27267&_BaoCaoThongKe_WAR_ctonegateportlet_mucDo=0&_BaoCaoThongKe_WAR_ctonegateportlet_cqqlid=610&_BaoCaoThongKe_WAR_ctonegateportlet_LoaiHS=0&_BaoCaoThongKe_WAR_ctonegateportlet_phongBanId=0&_BaoCaoThongKe_WAR_ctonegateportlet_phongBanArr=0&_BaoCaoThongKe_WAR_ctonegateportlet_loaiSoLieu=0"
        url_1cuank = f"https://motcua.ninhkieu.cantho.gov.vn/group/guest/bao-cao-thong-ke?p_p_id=BaoCaoThongKe_WAR_ctonegateportlet&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view&_BaoCaoThongKe_WAR_ctonegateportlet_jspPage=%2Fhtml%2Fonegate%2Fbaocaothongke%2Fhosodvclt%2Fthongkehoso.jsp"

        # Mở URL
        driver.get(url_1cuasotk)
        print("Page title: ", driver.title)
        
        # Đợi cho trang tải hoàn tất (trong trường hợp này, ví dụ chờ tiêu đề xuất hiện)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "title")))

        # Đăng nhập vào hệ thống
        login_id = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "_58_login")))
        login_id.send_keys('admin_stp')

        login_pwd = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "_58_password")))
        login_pwd.send_keys('Unitech@')

        login_pwd.send_keys(Keys.ENTER)

        # Chờ đợi cho trang tải hoàn tất sau khi đăng nhập
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//table[@class='tbl_baocao']"))
        )

        # Lấy giá trị trong cột thứ 4 của hàng 8 và cột thứ 6 của hàng 8
        row_8_col_4 = driver.find_element(By.XPATH, "//table[@class='tbl_baocao']//tr[8]/td[4]").text
        row_8_col_6 = driver.find_element(By.XPATH, "//table[@class='tbl_baocao']//tr[8]/td[6]").text

        # In ra các giá trị
        print("Thống kê ngày:", formatted_date)
        print("- Hồ sơ trực tuyến:", row_8_col_4)
        print("- Hồ sơ trực tiếp:", row_8_col_6)

        # Chuyển sang trang thống kê DVCLT
        driver.get(url_1cuank)

        # Đăng nhập vào hệ thống
        login_id_nk = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "_58_login")))
        login_id_nk.send_keys('admin_ninhkieu')

        login_pwd_nk = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "_58_password")))
        login_pwd_nk.send_keys('1cua@NK')

        login_pwd_nk.send_keys(Keys.ENTER)

        # Chờ trang tải xong trước khi tìm kiếm và điền ngày tháng
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "_BaoCaoThongKe_WAR_ctonegateportlet_tuNgay"))
        )

        # Điền ngày tháng vào trường 'từ ngày'
        tuNgay = driver.find_element(By.ID, "_BaoCaoThongKe_WAR_ctonegateportlet_tuNgay")
        tuNgay.clear()
        tuNgay.send_keys(formatted_yesterday)  # Chuỗi định dạng ngày hôm qua

        # Điền ngày tháng vào trường 'đến ngày'
        denNgay = driver.find_element(By.ID, "_BaoCaoThongKe_WAR_ctonegateportlet_denNgay")
        denNgay.clear()
        denNgay.send_keys(formatted_yesterday)  # Chuỗi định dạng ngày hôm qua

        # Nhấn nút xem
        xem = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "btnLuu")))
        xem.click()

        time.sleep(2)

        # Kiểm tra nếu có thông báo "Không có hồ sơ"
        try:
            ket_qua_div = driver.find_element(By.ID, "ketQua")
            khong_co_ho_so = ket_qua_div.find_element(By.CLASS_NAME, "alert-info").text == "Không có hồ sơ"
        except NoSuchElementException:
            khong_co_ho_so = False

        # Đường dẫn đến tệp Excel
        file_path = 'G:\\My Drive\\Thong_ke_DVCLT.xlsx'

        # Mở tệp Excel và chọn sheet đầu tiên
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        # Tìm hàng cuối cùng có dữ liệu
        last_row = find_last_row_with_data(sheet) + 1

        # Bổ sung giá trị vào các cột tương ứng
        if last_row > 2:
            previous_value = sheet[f'A{last_row - 1}'].value
            if previous_value is not None:
                sheet[f'A{last_row}'] = previous_value + 1  # Cột A tăng lên 1 so với hàng trước
            else:
                sheet[f'A{last_row}'] = 1  # Giá trị mặc định cho hàng đầu tiên
        else:
            sheet[f'A{last_row}'] = 1  # Giá trị mặc định cho hàng đầu tiên

        sheet[f'B{last_row}'] = formatted_date  # Ghi ngày vào cột B

        if khong_co_ho_so:
            sheet[f'F{last_row}'] = 0
            sheet[f'G{last_row}'] = 0
            sheet[f'H{last_row}'] = 0
            
        else:
            f_value = 0
            g_value = 0
            h_value = 0

            try:
                f_value_element = driver.find_element(By.XPATH, "//td[text()='Đăng ký khai tử, Xóa đăng ký thường trú']/following-sibling::td[1]")
                f_value = f_value_element.text
            except NoSuchElementException:
                pass

            try:
                g_value_element = driver.find_element(By.XPATH, "//td[text()='Liên thông thủ tục hành chính về đăng ký khai sinh, đăng ký thường trú, cấp thẻ bảo hiểm y tế cho trẻ em dưới 6 tuổi']/following-sibling::td[1]")
                g_value = g_value_element.text
            except NoSuchElementException:
                pass

            try:
                h_value_element = driver.find_element(By.XPATH, "//td[text()='Thủ tục liên thông về đăng ký khai tử, xóa đăng ký thường trú, hưởng chế độ tử tuất (trợ cấp tuất và trợ cấp mai táng)/hỗ trợ chi phí mai táng/hưởng mai táng phí']/following-sibling::td[1]")
                h_value = h_value_element.text
            except NoSuchElementException:
                pass

            sheet[f'F{last_row}'] = f_value if f_value is not None else 0
            sheet[f'G{last_row}'] = g_value if g_value is not None else 0
            sheet[f'H{last_row}'] = h_value if h_value is not None else 0

            # In ra các giá trị
            print("- Đăng ký khai tử, Xóa đăng ký thường trú:", f_value)
            print("- Đăng ký khai sinh, đăng ký thường trú, cấp thẻ bảo hiểm y tế cho trẻ em dưới 6 tuổi:", g_value)
            print("- Hỗ trợ chi phí mai táng/hưởng mai táng phí:", h_value)
        
        # Ghi giá trị vào các cột C và D
        sheet[f'C{last_row}'] = row_8_col_4
        sheet[f'D{last_row}'] = row_8_col_6

        # Tính tổng và ghi vào cột E
        sheet[f'E{last_row}'] = f'=C{last_row}+D{last_row}'

        # Canh giữa các ô trong các cột A, B, C, D, E, F, G, H
        for col in range(1, 9):  # Cột A đến H
            col_letter = get_column_letter(col)
            for row in range(3, last_row + 1):
                sheet[f'{col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center')

        # Vẽ viền all border từ cột A đến cột H cho hàng vừa điền dữ liệu
        for col in range(1, 9):  # Cột A đến H
            col_letter = get_column_letter(col)
            for row in range(last_row, last_row + 1):
                sheet[f'{col_letter}{row}'].border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Lưu tệp Excel
        workbook.save(filename=file_path)

        print(f"---> Đã ghi dữ liệu vào hàng {last_row}")

        # Chờ 1 giây trước khi đóng trình duyệt
        time.sleep(1)

    finally:
        # Đóng trình duyệt khi hoàn thành
        driver.quit()
        
    # Kết thúc chương trình
    os._exit(0)

if __name__ == "__main__":
    main()
